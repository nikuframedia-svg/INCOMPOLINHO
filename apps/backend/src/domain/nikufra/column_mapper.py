"""ISOP column detection and mapping logic.

Scans header rows to find column positions by pattern matching,
detects date columns, and extracts workday flags.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date

from openpyxl.worksheet.worksheet import Worksheet

from .helpers import _normalize_string, _parse_date_cell, _parse_numeric

# ── Column map ──────────────────────────────────────────────────


@dataclass
class ColumnMap:
    cliente: int = -1
    nome: int = -1
    ref_artigo: int = -1
    designacao: int = -1
    lote_econ: int = -1
    prz_fabrico: int = -1
    maquina: int = -1
    maq_alt: int = -1
    ferramenta: int = -1
    tp_setup: int = -1
    pecas_h: int = -1
    n_pessoas: int = -1
    qtd_exp: int = -1
    produto_acabado: int = -1
    stock_a: int = -1
    wip: int = -1
    atraso: int = -1
    estado_maq: int = -1
    estado_ferr: int = -1
    peca_gemea: int = -1


# ═══════════════════════════════════════════════════════════════
#  Header detection
# ═══════════════════════════════════════════════════════════════


def find_header_row(
    ws: Worksheet,
    max_scan: int = 16,
) -> tuple[int, list[str]] | None:
    """Scan rows 0-15 for the header row (must contain 'Referência Artigo' + 'Máquina')."""
    for ri in range(1, min(ws.max_row + 1, max_scan + 1)):
        row_cells = list(ws.iter_rows(min_row=ri, max_row=ri, values_only=True))[0]
        if len(row_cells) < 5:
            continue
        strs = [_normalize_string(c) for c in row_cells]
        has_ref = any(
            h in s
            for s in strs
            for h in (
                "Referência Artigo",
                "Referencia Artigo",
                "REFERÊNCIA ARTIGO",
                "Ref. Artigo",
                "REF. ARTIGO",
            )
        )
        has_maq = any(h in s for s in strs for h in ("Máquina", "Maquina", "MÁQUINA", "MAQUINA"))
        if has_ref and has_maq:
            return (ri, strs)
    return None


def build_column_map(headers: list[str]) -> ColumnMap | None:
    """Map column names to indices (0-based) based on header text patterns."""

    def find(*patterns: str) -> int:
        for i, h in enumerate(headers):
            hl = h.lower()
            for p in patterns:
                if p.lower() in hl:
                    return i
        return -1

    ref_artigo = find("referência artigo", "referencia artigo", "ref. artigo", "ref artigo")
    maquina = find("máquina", "maquina")

    if ref_artigo < 0 or maquina < 0:
        return None

    cm = ColumnMap()
    cm.ref_artigo = ref_artigo
    cm.maquina = maquina
    cm.cliente = find("cliente")
    cm.nome = find("nome")
    cm.designacao = find("designação", "designacao")
    cm.lote_econ = find("lote econ", "lote económico", "lote economico")
    cm.prz_fabrico = find("prz.fabrico", "prz fabrico", "prazo fabrico", "prazo de fabrico")
    cm.maq_alt = find("máq. alt", "maq. alt", "máquina alt", "maquina alt")
    cm.ferramenta = find("ferramenta")
    cm.tp_setup = find("tp.setup", "tp setup", "setup")
    cm.pecas_h = find(
        "peças/h",
        "pecas/h",
        "pcs/h",
        "pçs/h",
        "peças / h",
        "pecas / h",
        "cadência",
        "cadencia",
        "rate",
    )
    cm.n_pessoas = find(
        "nº pessoas",
        "n pessoas",
        "num pessoas",
        "nº pess",
        "n. pess",
        "pessoas",
    )
    cm.qtd_exp = find("qtd exp", "qtd. exp", "qtd expedição", "qtd expedicao")
    cm.produto_acabado = find("produto acabado", "prod. acabado", "prod acabado", "pa", "parent")
    cm.stock_a = find("stock-a", "stock a", "stock")
    cm.wip = find("wip")
    cm.atraso = find("atraso")
    cm.estado_maq = find(
        "estado máq",
        "estado maq",
        "status máq",
        "status maq",
        "estado máquina",
        "estado maquina",
    )
    cm.estado_ferr = find(
        "estado ferr",
        "status ferr",
        "estado ferramenta",
        "status ferramenta",
    )
    cm.peca_gemea = find("peca gemea", "peça gémea", "peça gemea", "pç gemea", "twin")
    return cm


def find_date_columns(
    header_cells: tuple,
    col_map: ColumnMap,
) -> tuple[list[date], list[int]]:
    """Find date columns in the header row.

    Returns:
        (dates, date_col_indices) — dates as date objects, indices as 0-based column numbers.
    """
    last_text_col = max(
        col_map.atraso,
        col_map.stock_a,
        col_map.wip,
        col_map.n_pessoas,
        col_map.pecas_h,
        col_map.ferramenta,
        col_map.qtd_exp,
        col_map.maquina,
    )
    date_search_start = last_text_col + 1 if last_text_col >= 0 else 10

    dates: list[date] = []
    date_col_indices: list[int] = []

    for ci in range(date_search_start, len(header_cells)):
        d = _parse_date_cell(header_cells[ci])
        if d:
            dates.append(d)
            date_col_indices.append(ci)

    # Fallback: search from column 10
    if not dates:
        for ci in range(10, len(header_cells)):
            d = _parse_date_cell(header_cells[ci])
            if d:
                dates.append(d)
                date_col_indices.append(ci)

    return dates, date_col_indices


def find_workday_flags_row(
    ws: Worksheet,
    header_row: int,
    date_col_indices: list[int],
) -> list[bool] | None:
    """Find workday flags row (0/1 values in date columns) above header row."""
    for ri in range(max(1, header_row - 4), header_row):
        row_vals = list(ws.iter_rows(min_row=ri, max_row=ri, values_only=True))[0]
        is_01 = True
        count = 0
        for ci in date_col_indices:
            if ci >= len(row_vals):
                continue
            v = row_vals[ci]
            if v is None:
                continue
            n = _parse_numeric(v, -1)
            if n != 0 and n != 1:
                is_01 = False
                break
            count += 1
        if is_01 and count > 3:
            return [
                _parse_numeric(
                    row_vals[ci] if ci < len(row_vals) else None,
                    1,
                )
                == 1
                for ci in date_col_indices
            ]
    return None
