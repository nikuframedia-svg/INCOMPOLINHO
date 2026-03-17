"""ISOP XLSX → NikufraData parser (Python port of frontend isop/ 923 LOC).

Mirrors the TypeScript parser logic:
  - Dynamic header detection (scan rows 0-15)
  - Flexible column mapping by pattern matching
  - Red cell highlighting detection (openpyxl styles)
  - Raw NP values in operations[].d
  - Trust score computation
  - Workday flag extraction

Output: NikufraData-compatible dict ready for scheduling pipeline.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from .constants import DAY_NAMES_PT

# ── Machine → Area mapping ──────────────────────────────────────

MACHINE_AREA: dict[str, str] = {
    "PRM019": "PG1",
    "PRM020": "PG1",
    "PRM043": "PG1",
    "PRM031": "PG2",
    "PRM039": "PG2",
    "PRM042": "PG2",
}


# ── Column map ──────────────────────────────────────────────────


@dataclass
class ColumnMap:
    cliente: int = -1
    nome: int = -1
    ref_artigo: int = -1
    designacao: int = -1
    lote_econ: int = -1
    prz_fabrico: int = -1
    maquina: int = -1
    maq_alt: int = -1
    ferramenta: int = -1
    tp_setup: int = -1
    pecas_h: int = -1
    n_pessoas: int = -1
    qtd_exp: int = -1
    produto_acabado: int = -1
    stock_a: int = -1
    wip: int = -1
    atraso: int = -1
    estado_maq: int = -1
    estado_ferr: int = -1
    peca_gemea: int = -1


# ── Parsed row ──────────────────────────────────────────────────


@dataclass
class ParsedRow:
    customer_code: str = ""
    customer_name: str = ""
    parent_sku: str = ""
    item_sku: str = ""
    item_name: str = ""
    lot_economic_qty: int = 0
    lead_time_days: int = 0
    resource_code: str = ""
    alt_resource: str = ""
    tool_code: str = ""
    setup_time: float = 0.0
    rate: float = 0.0
    operators_required: int = 1
    qtd_exp: float = 0.0
    stock: float = 0.0
    wip: float = 0.0
    atraso: float = 0.0
    daily_quantities: list[float | None] = field(default_factory=list)
    machine_down: bool = False
    tool_down: bool = False
    twin: str = ""


# ── Trust score result ──────────────────────────────────────────


@dataclass
class TrustScoreResult:
    score: float = 0.0
    dimensions: dict[str, float] = field(
        default_factory=lambda: {
            "completeness": 0.0,
            "quality": 0.0,
            "demandCoverage": 0.0,
            "consistency": 0.0,
        }
    )


# ── Parse result ────────────────────────────────────────────────


@dataclass
class ISOPParseResult:
    success: bool = True
    data: dict[str, Any] | None = None  # NikufraData dict
    meta: dict[str, Any] | None = None
    source_columns: dict[str, bool] | None = None
    errors: list[str] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════════


def _normalize_code(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def _normalize_string(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_numeric(value: Any, fallback: float = 0.0) -> float:
    if value is None:
        return fallback
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", ".").strip()
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return fallback
    return fallback


def _parse_integer(value: Any, fallback: int = 1) -> int:
    return round(_parse_numeric(value, float(fallback)))


def _format_date(d: date) -> str:
    return d.strftime("%d/%m")


def _day_label(d: date) -> str:
    return DAY_NAMES_PT[d.weekday()]


def _parse_date_cell(value: Any) -> date | None:
    """Parse a date cell from Excel — handles datetime, date, numeric serial, strings."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date: days since 1899-12-30
        try:
            serial = int(value)
            if serial < 1 or serial > 200000:
                return None
            base = datetime(1899, 12, 30)
            from datetime import timedelta

            return (base + timedelta(days=serial)).date()
        except (ValueError, OverflowError):
            return None
    if isinstance(value, str):
        s = value.strip()
        # ISO: YYYY-MM-DD
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
        if m:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        # DD/MM/YYYY
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
        if m:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        # DD/MM/YY
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2})$", s)
        if m:
            yr = int(m.group(3))
            yr = yr + 2000 if yr < 70 else yr + 1900
            return date(yr, int(m.group(2)), int(m.group(1)))
        # DD/MM (assume current year)
        m = re.match(r"^(\d{1,2})/(\d{1,2})$", s)
        if m:
            return date(date.today().year, int(m.group(2)), int(m.group(1)))
    return None


# ── Red cell detection ──────────────────────────────────────────


def _is_red_color(rgb: str | None) -> bool:
    """Detect red tones: high R, low G and B."""
    if not rgb or not isinstance(rgb, str):
        return False
    hex_str = rgb[-6:] if len(rgb) >= 6 else rgb
    if len(hex_str) != 6:
        return False
    try:
        r = int(hex_str[0:2], 16)
        g = int(hex_str[2:4], 16)
        b = int(hex_str[4:6], 16)
        return r > 180 and g < 100 and b < 100
    except ValueError:
        return False


def _is_cell_red_highlighted(cell: Cell) -> bool:
    """Check if a cell has red fill or red font color."""
    try:
        # Check fill color
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            rgb = str(cell.fill.fgColor.rgb)
            if _is_red_color(rgb):
                return True
        if cell.fill and cell.fill.bgColor and cell.fill.bgColor.rgb:
            rgb = str(cell.fill.bgColor.rgb)
            if _is_red_color(rgb):
                return True
        # Check font color
        if cell.font and cell.font.color and cell.font.color.rgb:
            rgb = str(cell.font.color.rgb)
            if _is_red_color(rgb):
                return True
    except (AttributeError, TypeError):
        pass
    return False


# ═══════════════════════════════════════════════════════════════
#  Header detection
# ═══════════════════════════════════════════════════════════════


def _find_header_row(
    ws: Worksheet,
    max_scan: int = 16,
) -> tuple[int, list[str]] | None:
    """Scan rows 0-15 for the header row (must contain 'Referência Artigo' + 'Máquina')."""
    for ri in range(1, min(ws.max_row + 1, max_scan + 1)):
        row_cells = list(ws.iter_rows(min_row=ri, max_row=ri, values_only=True))[0]
        if len(row_cells) < 5:
            continue
        strs = [_normalize_string(c) for c in row_cells]
        has_ref = any(
            h in s
            for s in strs
            for h in (
                "Referência Artigo",
                "Referencia Artigo",
                "REFERÊNCIA ARTIGO",
                "Ref. Artigo",
                "REF. ARTIGO",
            )
        )
        has_maq = any(h in s for s in strs for h in ("Máquina", "Maquina", "MÁQUINA", "MAQUINA"))
        if has_ref and has_maq:
            return (ri, strs)
    return None


def _build_column_map(headers: list[str]) -> ColumnMap | None:
    """Map column names to indices (0-based) based on header text patterns."""

    def find(*patterns: str) -> int:
        for i, h in enumerate(headers):
            hl = h.lower()
            for p in patterns:
                if p.lower() in hl:
                    return i
        return -1

    ref_artigo = find("referência artigo", "referencia artigo", "ref. artigo", "ref artigo")
    maquina = find("máquina", "maquina")

    if ref_artigo < 0 or maquina < 0:
        return None

    cm = ColumnMap()
    cm.ref_artigo = ref_artigo
    cm.maquina = maquina
    cm.cliente = find("cliente")
    cm.nome = find("nome")
    cm.designacao = find("designação", "designacao")
    cm.lote_econ = find("lote econ", "lote económico", "lote economico")
    cm.prz_fabrico = find("prz.fabrico", "prz fabrico", "prazo fabrico", "prazo de fabrico")
    cm.maq_alt = find("máq. alt", "maq. alt", "máquina alt", "maquina alt")
    cm.ferramenta = find("ferramenta")
    cm.tp_setup = find("tp.setup", "tp setup", "setup")
    cm.pecas_h = find(
        "peças/h",
        "pecas/h",
        "pcs/h",
        "pçs/h",
        "peças / h",
        "pecas / h",
        "cadência",
        "cadencia",
        "rate",
    )
    cm.n_pessoas = find(
        "nº pessoas",
        "n pessoas",
        "num pessoas",
        "nº pess",
        "n. pess",
        "pessoas",
    )
    cm.qtd_exp = find("qtd exp", "qtd. exp", "qtd expedição", "qtd expedicao")
    cm.produto_acabado = find("produto acabado", "prod. acabado", "prod acabado", "pa", "parent")
    cm.stock_a = find("stock-a", "stock a", "stock")
    cm.wip = find("wip")
    cm.atraso = find("atraso")
    cm.estado_maq = find(
        "estado máq",
        "estado maq",
        "status máq",
        "status maq",
        "estado máquina",
        "estado maquina",
    )
    cm.estado_ferr = find(
        "estado ferr",
        "status ferr",
        "estado ferramenta",
        "status ferramenta",
    )
    cm.peca_gemea = find("peca gemea", "peça gémea", "peça gemea", "pç gemea", "twin")
    return cm


def _find_workday_flags_row(
    ws: Worksheet,
    header_row: int,
    date_col_indices: list[int],
) -> list[bool] | None:
    """Find workday flags row (0/1 values in date columns) above header row."""
    for ri in range(max(1, header_row - 4), header_row):
        row_vals = list(ws.iter_rows(min_row=ri, max_row=ri, values_only=True))[0]
        is_01 = True
        count = 0
        for ci in date_col_indices:
            if ci >= len(row_vals):
                continue
            v = row_vals[ci]
            if v is None:
                continue
            n = _parse_numeric(v, -1)
            if n != 0 and n != 1:
                is_01 = False
                break
            count += 1
        if is_01 and count > 3:
            return [
                _parse_numeric(
                    row_vals[ci] if ci < len(row_vals) else None,
                    1,
                )
                == 1
                for ci in date_col_indices
            ]
    return None


# ═══════════════════════════════════════════════════════════════
#  Trust score
# ═══════════════════════════════════════════════════════════════


def _compute_trust_score(
    rows: list[ParsedRow],
    tools: list[dict[str, Any]],
    operations: list[dict[str, Any]],
) -> TrustScoreResult:
    if not rows:
        return TrustScoreResult()

    n = len(rows)

    # 1. Completeness (40%)
    complete = sum(
        1
        for r in rows
        if r.item_sku and r.resource_code and r.tool_code and r.rate > 0 and r.setup_time >= 0
    )
    completeness = complete / n

    # 2. Quality (30%)
    valid = sum(1 for r in rows if r.rate > 0 and r.setup_time >= 0 and r.operators_required >= 1)
    quality = valid / n

    # 3. Demand coverage (20%)
    with_demand = sum(
        1 for op in operations if any(v is not None and v != 0 for v in op.get("d", []))
    )
    demand_coverage = with_demand / len(operations) if operations else 0.0

    # 4. Consistency (10%)
    machine_set = {r.resource_code for r in rows}
    valid_tools = sum(1 for t in tools if t.get("m") in machine_set)
    consistency = valid_tools / len(tools) if tools else 1.0

    score = round(
        completeness * 0.4 + quality * 0.3 + demand_coverage * 0.2 + consistency * 0.1,
        2,
    )

    return TrustScoreResult(
        score=score,
        dimensions={
            "completeness": round(completeness, 2),
            "quality": round(quality, 2),
            "demandCoverage": round(demand_coverage, 2),
            "consistency": round(consistency, 2),
        },
    )


# ═══════════════════════════════════════════════════════════════
#  Build NikufraData
# ═══════════════════════════════════════════════════════════════


def _build_nikufra_data(
    parsed_rows: list[ParsedRow],
    dates: list[date],
    workday_flags: list[bool],
    warnings: list[str],
    source_columns: dict[str, bool],
) -> ISOPParseResult:
    n_days = len(dates)

    # ── Machines ──
    machine_set: dict[str, str] = {}  # id → area
    for r in parsed_rows:
        if r.resource_code not in machine_set:
            machine_set[r.resource_code] = MACHINE_AREA.get(r.resource_code, "PG1")
        if r.alt_resource and r.alt_resource not in machine_set:
            machine_set[r.alt_resource] = MACHINE_AREA.get(r.alt_resource, "PG1")

    machines_down = {r.resource_code for r in parsed_rows if r.machine_down}
    if machines_down:
        warnings.append(
            "Máquinas inoperacionais detectadas (texto/cor): " + ", ".join(sorted(machines_down))
        )

    unknown_machines = [mid for mid in machine_set if mid not in MACHINE_AREA]
    if unknown_machines:
        warnings.append(
            "Máquina(s) desconhecida(s) atribuída(s) a PG1 por defeito: "
            + ", ".join(unknown_machines)
        )

    machines = []
    for mid in sorted(machine_set):
        m: dict[str, Any] = {
            "id": mid,
            "area": machine_set[mid],
            "man": [0] * n_days,
        }
        if mid in machines_down:
            m["status"] = "down"
        machines.append(m)

    # ── Tools ──
    tool_map: dict[str, dict[str, Any]] = {}
    for row in parsed_rows:
        if not row.tool_code:
            continue
        existing = tool_map.get(row.tool_code)
        if existing:
            if row.item_sku not in existing["skus"]:
                existing["skus"].append(row.item_sku)
                existing["nm"].append(row.item_name)
            if row.resource_code != existing["m"]:
                warnings.append(
                    f'Ferramenta "{row.tool_code}" aparece com máquinas diferentes: '
                    f"{existing['m']} (mantida) vs {row.resource_code} (SKU {row.item_sku})"
                )
            existing["wip"] = max(existing.get("wip", 0), row.wip)
        else:
            tool_map[row.tool_code] = {
                "id": row.tool_code,
                "m": row.resource_code,
                "alt": row.alt_resource or "-",
                "s": row.setup_time,
                "pH": row.rate,
                "op": row.operators_required,
                "skus": [row.item_sku],
                "nm": [row.item_name],
                "lt": row.lot_economic_qty,
                "stk": 0,
                "wip": row.wip,
            }

    tools_down = {r.tool_code for r in parsed_rows if r.tool_down and r.tool_code}
    if tools_down:
        warnings.append(
            "Ferramentas inoperacionais detectadas (texto/cor): " + ", ".join(sorted(tools_down))
        )

    tools = list(tool_map.values())
    for t in tools:
        if t["id"] in tools_down:
            t["status"] = "down"

    # ── Customers ──
    customer_map: dict[str, str] = {}
    for r in parsed_rows:
        if r.customer_code and r.customer_code not in customer_map:
            customer_map[r.customer_code] = r.customer_name
    customers = [{"code": code, "name": name} for code, name in sorted(customer_map.items())]

    # ── Operations ──
    ops_without_tool = [r for r in parsed_rows if not r.tool_code]
    if ops_without_tool:
        skus = ", ".join(r.item_sku for r in ops_without_tool[:5])
        suffix = "…" if len(ops_without_tool) > 5 else ""
        warnings.append(
            f"{len(ops_without_tool)} operação(ões) sem código de ferramenta — "
            f"não serão agendadas: {skus}{suffix}"
        )

    operations: list[dict[str, Any]] = []
    for idx, row in enumerate(parsed_rows):
        op: dict[str, Any] = {
            "id": f"OP{idx + 1:02d}",
            "m": row.resource_code,
            "t": row.tool_code,
            "sku": row.item_sku,
            "nm": row.item_name,
            "pH": row.rate,
            "atr": row.atraso,
            "d": row.daily_quantities,
            "s": row.setup_time,
            "op": row.operators_required,
        }
        if row.customer_code:
            op["cl"] = row.customer_code
        if row.customer_name:
            op["clNm"] = row.customer_name
        if row.parent_sku:
            op["pa"] = row.parent_sku
        if row.wip:
            op["wip"] = row.wip
        if row.qtd_exp:
            op["qe"] = row.qtd_exp
        if row.lead_time_days:
            op["ltDays"] = row.lead_time_days
        if row.twin:
            op["twin"] = row.twin
        operations.append(op)

    # ── MO load ──
    mo = {"PG1": [0.0] * n_days, "PG2": [0.0] * n_days}

    # ── Date labels ──
    date_labels = [_format_date(d) for d in dates]
    day_labels = [_day_label(d) for d in dates]

    nikufra_data: dict[str, Any] = {
        "dates": date_labels,
        "days_label": day_labels,
        "mo": mo,
        "machines": machines,
        "tools": tools,
        "operations": operations,
        "history": [],
        "customers": customers,
        "workday_flags": workday_flags,
    }

    # ── Trust score ──
    trust = _compute_trust_score(parsed_rows, tools, operations)

    unique_skus = {r.item_sku for r in parsed_rows}
    unique_machines = {r.resource_code for r in parsed_rows}
    unique_tools = {r.tool_code for r in parsed_rows if r.tool_code}
    workday_count = sum(1 for f in workday_flags if f)

    meta: dict[str, Any] = {
        "rows": len(parsed_rows),
        "machines": len(unique_machines),
        "tools": len(unique_tools),
        "skus": len(unique_skus),
        "dates": n_days,
        "workdays": workday_count,
        "trustScore": trust.score,
        "trustDimensions": trust.dimensions,
        "warnings": warnings,
    }

    missing = [k for k, v in source_columns.items() if not v]
    if missing:
        warnings.append(
            "Colunas não detectadas: "
            + ", ".join(missing)
            + " — serão preenchidas pelo ISOP Mestre ou defaults."
        )

    return ISOPParseResult(
        success=True,
        data=nikufra_data,
        meta=meta,
        source_columns=source_columns,
    )


# ═══════════════════════════════════════════════════════════════
#  Main parser
# ═══════════════════════════════════════════════════════════════

_DOWN_PATTERN = re.compile(r"inact|down|avaria|parad|inoper", re.IGNORECASE)


def parse_isop_file(
    filepath_or_bytes: Any,
    *,
    data_only: bool = True,
) -> ISOPParseResult:
    """Parse an ISOP XLSX file and return NikufraData.

    Args:
        filepath_or_bytes: Path string, pathlib.Path, or bytes/BytesIO.
        data_only: If True, read cached values (not formulas).

    Returns:
        ISOPParseResult with success flag, data (NikufraData dict), meta, errors.
    """
    # 1. Load workbook
    try:
        wb = openpyxl.load_workbook(filepath_or_bytes, data_only=data_only)
    except Exception:
        return ISOPParseResult(
            success=False,
            errors=["Ficheiro XLSX inválido — não foi possível abrir."],
        )

    # 2. Find sheet "Planilha1"
    sheet_name: str | None = None
    for sn in wb.sheetnames:
        if sn.lower() == "planilha1":
            sheet_name = sn
            break

    if not sheet_name:
        return ISOPParseResult(
            success=False,
            errors=['Sheet "Planilha1" não encontrada no ficheiro.'],
        )

    ws: Worksheet = wb[sheet_name]

    # 3. Find header row dynamically
    header_result = _find_header_row(ws)
    if not header_result:
        return ISOPParseResult(
            success=False,
            errors=[
                'Nenhuma linha de cabeçalho encontrada com "Referência Artigo" e "Máquina" '
                "(procurado nas primeiras 16 linhas)."
            ],
        )

    header_row_idx, headers = header_result  # 1-based row number

    # 4. Map columns by header name
    col_map = _build_column_map(headers)
    if not col_map:
        return ISOPParseResult(
            success=False,
            errors=['Colunas "Referência Artigo" e "Máquina" não encontradas nos cabeçalhos.'],
        )

    # 5. Find date columns (after last text column)
    header_cells = list(
        ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)
    )[0]

    last_text_col = max(
        col_map.atraso,
        col_map.stock_a,
        col_map.wip,
        col_map.n_pessoas,
        col_map.pecas_h,
        col_map.ferramenta,
        col_map.qtd_exp,
        col_map.maquina,
    )
    date_search_start = last_text_col + 1 if last_text_col >= 0 else 10

    dates: list[date] = []
    date_col_indices: list[int] = []  # 0-based indices

    for ci in range(date_search_start, len(header_cells)):
        d = _parse_date_cell(header_cells[ci])
        if d:
            dates.append(d)
            date_col_indices.append(ci)

    # Fallback: search from column 10
    if not dates:
        for ci in range(10, len(header_cells)):
            d = _parse_date_cell(header_cells[ci])
            if d:
                dates.append(d)
                date_col_indices.append(ci)

    if not dates:
        return ISOPParseResult(
            success=False,
            errors=[f"Nenhuma coluna de data encontrada no cabeçalho (linha {header_row_idx})."],
        )

    # 6. Parse workday flags
    workday_flags = _find_workday_flags_row(ws, header_row_idx, date_col_indices)
    if workday_flags is None:
        workday_flags = [d.weekday() < 5 for d in dates]
    if len(workday_flags) != len(dates):
        workday_flags = [True] * len(dates)

    # 7. Parse data rows
    data_start_row = header_row_idx + 1  # 1-based
    parsed_rows: list[ParsedRow] = []
    warnings: list[str] = []

    for ri in range(data_start_row, ws.max_row + 1):
        row_cells = [ws.cell(row=ri, column=c + 1) for c in range(ws.max_column)]
        row_vals = [c.value for c in row_cells]

        if not row_vals or all(v is None for v in row_vals):
            continue

        item_sku = _normalize_code(
            row_vals[col_map.ref_artigo] if col_map.ref_artigo < len(row_vals) else None
        )
        if not item_sku:
            continue

        resource_code = _normalize_code(
            row_vals[col_map.maquina] if col_map.maquina < len(row_vals) else None
        )
        if not resource_code:
            warnings.append(f'Linha {ri}: SKU "{item_sku}" sem máquina — ignorada.')
            continue

        def _val(idx: int) -> Any:
            if idx < 0 or idx >= len(row_vals):
                return None
            return row_vals[idx]

        def _cell(idx: int) -> Cell | None:
            if idx < 0 or idx >= len(row_cells):
                return None
            return row_cells[idx]

        alt_resource = _normalize_code(_val(col_map.maq_alt))
        tool_code = _normalize_code(_val(col_map.ferramenta))
        rate = _parse_numeric(_val(col_map.pecas_h)) if col_map.pecas_h >= 0 else 0.0

        if rate <= 0 and col_map.pecas_h >= 0:
            warnings.append(
                f'Linha {ri}: SKU "{item_sku}" rate=0 — incluída mas não será agendada.'
            )

        # Parse daily NP values + red cell detection
        invalid_cells = 0
        daily_quantities: list[float | None] = []
        for ci in date_col_indices:
            cell = _cell(ci)
            raw = cell.value if cell else None
            if raw is None or (isinstance(raw, str) and raw.strip() == ""):
                daily_quantities.append(None)
                continue
            if isinstance(raw, str):
                try:
                    float(raw.replace(",", "."))
                except ValueError:
                    invalid_cells += 1
                    daily_quantities.append(None)
                    continue
            val = _parse_numeric(raw)
            # Red font/fill = demand (positive displayed red → negative NP)
            if val > 0 and cell is not None and _is_cell_red_highlighted(cell):
                val = -val
            daily_quantities.append(val)

        if invalid_cells > 0:
            warnings.append(
                f'Linha {ri}: SKU "{item_sku}" tem {invalid_cells} célula(s) de data '
                "não-numérica(s) — interpretada(s) como 0."
            )

        # Machine/tool down detection
        machine_down = False
        if col_map.estado_maq >= 0:
            estado = _normalize_string(_val(col_map.estado_maq))
            if _DOWN_PATTERN.search(estado):
                machine_down = True
        maq_cell = _cell(col_map.maquina)
        if maq_cell and _is_cell_red_highlighted(maq_cell):
            machine_down = True

        tool_down = False
        if col_map.estado_ferr >= 0:
            estado = _normalize_string(_val(col_map.estado_ferr))
            if _DOWN_PATTERN.search(estado):
                tool_down = True
        if col_map.ferramenta >= 0:
            ferr_cell = _cell(col_map.ferramenta)
            if ferr_cell and _is_cell_red_highlighted(ferr_cell):
                tool_down = True

        twin = _normalize_code(_val(col_map.peca_gemea))

        parsed_rows.append(
            ParsedRow(
                customer_code=_normalize_code(_val(col_map.cliente))
                if col_map.cliente >= 0
                else "",
                customer_name=_normalize_string(_val(col_map.nome)) if col_map.nome >= 0 else "",
                parent_sku=_normalize_code(_val(col_map.produto_acabado))
                if col_map.produto_acabado >= 0
                else "",
                item_sku=item_sku,
                item_name=_normalize_string(_val(col_map.designacao))
                if col_map.designacao >= 0
                else item_sku,
                lot_economic_qty=_parse_integer(_val(col_map.lote_econ), 0)
                if col_map.lote_econ >= 0
                else 0,
                lead_time_days=_parse_integer(_val(col_map.prz_fabrico), 0)
                if col_map.prz_fabrico >= 0
                else 0,
                resource_code=resource_code,
                alt_resource="" if alt_resource == "-" else alt_resource,
                tool_code=tool_code,
                setup_time=_parse_numeric(_val(col_map.tp_setup)) if col_map.tp_setup >= 0 else 0.0,
                rate=rate,
                operators_required=_parse_integer(_val(col_map.n_pessoas), 1)
                if col_map.n_pessoas >= 0
                else 1,
                qtd_exp=_parse_numeric(_val(col_map.qtd_exp)) if col_map.qtd_exp >= 0 else 0.0,
                stock=0.0,
                wip=_parse_numeric(_val(col_map.wip)) if col_map.wip >= 0 else 0.0,
                atraso=_parse_numeric(_val(col_map.atraso)) if col_map.atraso >= 0 else 0.0,
                daily_quantities=daily_quantities,
                machine_down=machine_down,
                tool_down=tool_down,
                twin=twin,
            )
        )

    if not parsed_rows:
        return ISOPParseResult(
            success=False,
            errors=[
                f"Nenhuma linha de dados válida encontrada (a partir da linha {data_start_row})."
            ],
        )

    warnings.append(
        f"Cabeçalho detectado na linha {header_row_idx}, "
        f"dados a partir da linha {data_start_row}, "
        f"{len(dates)} datas, {len(parsed_rows)} operações."
    )

    # 8. Build NikufraData
    source_columns = {
        "hasSetup": col_map.tp_setup >= 0,
        "hasAltMachine": col_map.maq_alt >= 0,
        "hasRate": col_map.pecas_h >= 0,
        "hasParentSku": col_map.produto_acabado >= 0,
        "hasLeadTime": col_map.prz_fabrico >= 0,
        "hasQtdExp": col_map.qtd_exp >= 0,
        "hasTwin": col_map.peca_gemea >= 0,
    }

    return _build_nikufra_data(parsed_rows, dates, workday_flags, warnings, source_columns)
