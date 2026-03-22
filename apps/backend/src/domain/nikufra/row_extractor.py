"""ISOP row parsing and extraction logic.

Extracts ParsedRow objects from worksheet data rows using the column map.
Handles NP value parsing, red cell detection, and machine/tool down status.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from .column_mapper import ColumnMap
from .helpers import (
    _is_cell_red_highlighted,
    _normalize_code,
    _normalize_string,
    _parse_integer,
    _parse_numeric,
)

# ── Parsed row ──────────────────────────────────────────────────


@dataclass
class ParsedRow:
    customer_code: str = ""
    customer_name: str = ""
    parent_sku: str = ""
    item_sku: str = ""
    item_name: str = ""
    lot_economic_qty: int = 0
    lead_time_days: int = 0
    resource_code: str = ""
    alt_resource: str = ""
    tool_code: str = ""
    setup_time: float = 0.0
    rate: float = 0.0
    operators_required: int = 1
    qtd_exp: float = 0.0
    stock: float = 0.0
    wip: float = 0.0
    atraso: float = 0.0
    daily_quantities: list[float | None] = field(default_factory=list)
    machine_down: bool = False
    tool_down: bool = False
    twin: str = ""


# ── Down detection pattern ──────────────────────────────────────

_DOWN_PATTERN = re.compile(r"inact|down|avaria|parad|inoper", re.IGNORECASE)


# ═══════════════════════════════════════════════════════════════
#  Row extraction
# ═══════════════════════════════════════════════════════════════


def extract_rows(
    ws: Worksheet,
    col_map: ColumnMap,
    date_col_indices: list[int],
    data_start_row: int,
) -> tuple[list[ParsedRow], list[str]]:
    """Parse all data rows from the worksheet.

    Args:
        ws: The openpyxl worksheet.
        col_map: Column index mapping.
        date_col_indices: 0-based column indices for date columns.
        data_start_row: 1-based row number where data starts.

    Returns:
        (parsed_rows, warnings) tuple.
    """
    parsed_rows: list[ParsedRow] = []
    warnings: list[str] = []

    for ri in range(data_start_row, ws.max_row + 1):
        row_cells = [ws.cell(row=ri, column=c + 1) for c in range(ws.max_column)]
        row_vals = [c.value for c in row_cells]

        if not row_vals or all(v is None for v in row_vals):
            continue

        item_sku = _normalize_code(
            row_vals[col_map.ref_artigo] if col_map.ref_artigo < len(row_vals) else None
        )
        if not item_sku:
            continue

        resource_code = _normalize_code(
            row_vals[col_map.maquina] if col_map.maquina < len(row_vals) else None
        )
        if not resource_code:
            warnings.append(f'Linha {ri}: SKU "{item_sku}" sem máquina — ignorada.')
            continue

        def _val(idx: int) -> Any:
            if idx < 0 or idx >= len(row_vals):
                return None
            return row_vals[idx]

        def _cell(idx: int) -> Cell | None:
            if idx < 0 or idx >= len(row_cells):
                return None
            return row_cells[idx]

        alt_resource = _normalize_code(_val(col_map.maq_alt))
        tool_code = _normalize_code(_val(col_map.ferramenta))
        rate = _parse_numeric(_val(col_map.pecas_h)) if col_map.pecas_h >= 0 else 0.0

        if rate <= 0 and col_map.pecas_h >= 0:
            warnings.append(
                f'Linha {ri}: SKU "{item_sku}" rate=0 — incluída mas não será agendada.'
            )

        # Parse daily NP values + red cell detection
        daily_quantities = _parse_daily_quantities(
            row_cells, date_col_indices, ri, item_sku, warnings
        )

        # Machine/tool down detection
        machine_down = _detect_machine_down(col_map, _val, _cell)
        tool_down = _detect_tool_down(col_map, _val, _cell)

        twin = _normalize_code(_val(col_map.peca_gemea))

        parsed_rows.append(
            ParsedRow(
                customer_code=_normalize_code(_val(col_map.cliente))
                if col_map.cliente >= 0
                else "",
                customer_name=_normalize_string(_val(col_map.nome)) if col_map.nome >= 0 else "",
                parent_sku=_normalize_code(_val(col_map.produto_acabado))
                if col_map.produto_acabado >= 0
                else "",
                item_sku=item_sku,
                item_name=_normalize_string(_val(col_map.designacao))
                if col_map.designacao >= 0
                else item_sku,
                lot_economic_qty=_parse_integer(_val(col_map.lote_econ), 0)
                if col_map.lote_econ >= 0
                else 0,
                lead_time_days=_parse_integer(_val(col_map.prz_fabrico), 0)
                if col_map.prz_fabrico >= 0
                else 0,
                resource_code=resource_code,
                alt_resource="" if alt_resource == "-" else alt_resource,
                tool_code=tool_code,
                setup_time=_parse_numeric(_val(col_map.tp_setup)) if col_map.tp_setup >= 0 else 0.0,
                rate=rate,
                operators_required=_parse_integer(_val(col_map.n_pessoas), 1)
                if col_map.n_pessoas >= 0
                else 1,
                qtd_exp=_parse_numeric(_val(col_map.qtd_exp)) if col_map.qtd_exp >= 0 else 0.0,
                stock=0.0,
                wip=_parse_numeric(_val(col_map.wip)) if col_map.wip >= 0 else 0.0,
                atraso=_parse_numeric(_val(col_map.atraso)) if col_map.atraso >= 0 else 0.0,
                daily_quantities=daily_quantities,
                machine_down=machine_down,
                tool_down=tool_down,
                twin=twin,
            )
        )

    return parsed_rows, warnings


# ── Internal helpers ────────────────────────────────────────────


def _parse_daily_quantities(
    row_cells: list[Cell],
    date_col_indices: list[int],
    row_index: int,
    item_sku: str,
    warnings: list[str],
) -> list[float | None]:
    """Parse daily NP values from date columns, handling red cell detection."""
    invalid_cells = 0
    daily_quantities: list[float | None] = []

    for ci in date_col_indices:
        cell = row_cells[ci] if ci < len(row_cells) else None
        raw = cell.value if cell else None
        if raw is None or (isinstance(raw, str) and raw.strip() == ""):
            daily_quantities.append(None)
            continue
        if isinstance(raw, str):
            try:
                float(raw.replace(",", "."))
            except ValueError:
                invalid_cells += 1
                daily_quantities.append(None)
                continue
        val = _parse_numeric(raw)
        # Red font/fill = demand (positive displayed red -> negative NP)
        if val > 0 and cell is not None and _is_cell_red_highlighted(cell):
            val = -val
        daily_quantities.append(val)

    if invalid_cells > 0:
        warnings.append(
            f'Linha {row_index}: SKU "{item_sku}" tem {invalid_cells} célula(s) de data '
            "não-numérica(s) — interpretada(s) como 0."
        )

    return daily_quantities


def _detect_machine_down(
    col_map: ColumnMap,
    _val: Any,
    _cell: Any,
) -> bool:
    """Detect if machine is down from status column or red highlighting."""
    machine_down = False
    if col_map.estado_maq >= 0:
        estado = _normalize_string(_val(col_map.estado_maq))
        if _DOWN_PATTERN.search(estado):
            machine_down = True
    maq_cell = _cell(col_map.maquina)
    if maq_cell and _is_cell_red_highlighted(maq_cell):
        machine_down = True
    return machine_down


def _detect_tool_down(
    col_map: ColumnMap,
    _val: Any,
    _cell: Any,
) -> bool:
    """Detect if tool is down from status column or red highlighting."""
    tool_down = False
    if col_map.estado_ferr >= 0:
        estado = _normalize_string(_val(col_map.estado_ferr))
        if _DOWN_PATTERN.search(estado):
            tool_down = True
    if col_map.ferramenta >= 0:
        ferr_cell = _cell(col_map.ferramenta)
        if ferr_cell and _is_cell_red_highlighted(ferr_cell):
            tool_down = True
    return tool_down
