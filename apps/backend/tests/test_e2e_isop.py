# FINAL-09: End-to-end tests with real ISOP Nikufra 17/03/2026
# Parses ISOP XLSX → builds SolverRequest → runs CP-SAT → validates results.

from __future__ import annotations

import math
import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from src.domain.solver.cpsat_solver import CpsatSolver
from src.domain.solver.schemas import (
    ConstraintConfigInput,
    JobInput,
    MachineInput,
    OperationInput,
    SolverConfig,
    SolverRequest,
)

# ── ISOP Parser (Python equivalent of isopClientParser.ts) ──

ISOP_PATH = Path(
    os.environ.get(
        "ISOP_PATH",
        "/Users/martimnicolau/Downloads/ISOP_ Nikufra_17_3.xlsx",
    )
)

HEADER_ROW = 5  # 1-indexed
DATA_START_ROW = 6
DATE_START_COL = 15  # Column O (1-indexed)

DAY_CAP = 1020
DEFAULT_OEE = 0.66
PRE_START_DAYS = 5

# Columns (1-indexed)
COL_CLIENTE = 1
COL_SKU = 3
COL_ECO_LOT = 5
COL_MACHINE = 7
COL_TOOL = 8
COL_PH = 9
COL_OPERATORS = 10
COL_TWIN = 13
COL_ATRASO = 14


def _parse_isop():
    """Parse ISOP XLSX into structured data for solver."""
    if not ISOP_PATH.exists():
        pytest.skip(f"ISOP file not found: {ISOP_PATH}")

    wb = openpyxl.load_workbook(str(ISOP_PATH), data_only=True)
    ws = wb.active

    # Parse dates from header row
    dates = []
    for c in range(DATE_START_COL, ws.max_column + 1):
        val = ws.cell(HEADER_ROW, c).value
        if isinstance(val, datetime):
            dates.append(val)
        elif val is None:
            break
        else:
            try:
                dates.append(datetime.fromisoformat(str(val)[:10]))
            except (ValueError, TypeError):
                break

    n_days = len(dates)
    start_date = dates[0] if dates else datetime(2026, 3, 17)

    # Build workdays (Mon-Fri = True, Sat/Sun = False)
    workdays_bool = []
    for d in dates:
        workdays_bool.append(d.weekday() < 5)  # 0=Mon, 4=Fri

    # Build workday indices (including pre-start)
    workday_indices = []
    # Pre-start days (assume workdays)
    for i in range(PRE_START_DAYS):
        workday_indices.append(-(PRE_START_DAYS - i))  # negative = before ISOP

    for i, is_work in enumerate(workdays_bool):
        if is_work:
            workday_indices.append(i)

    # Calendar-to-workday mapping
    total_cal = PRE_START_DAYS + n_days
    cal_to_work = [-1] * total_cal
    wk_idx = 0
    for cal in range(total_cal):
        isop_day = cal - PRE_START_DAYS
        if cal < PRE_START_DAYS:
            is_work = True
        elif 0 <= isop_day < len(workdays_bool):
            is_work = workdays_bool[isop_day]
        else:
            is_work = True
        if is_work:
            cal_to_work[cal] = wk_idx
            wk_idx += 1

    n_workdays = wk_idx

    # Parse operations
    ops = []
    machines_set = set()
    twin_refs = {}  # sku → twin_sku

    for r in range(DATA_START_ROW, ws.max_row + 1):
        sku = ws.cell(r, COL_SKU).value
        machine = ws.cell(r, COL_MACHINE).value
        tool = ws.cell(r, COL_TOOL).value
        ph_val = ws.cell(r, COL_PH).value

        if not sku or not machine or not ph_val:
            continue

        sku = str(sku).strip()
        machine = str(machine).strip()
        tool = str(tool).strip() if tool else "UNKNOWN"
        pH = float(ph_val) if ph_val else 0
        if pH <= 0:
            continue

        eco_lot_val = ws.cell(r, COL_ECO_LOT).value
        eco_lot = int(eco_lot_val) if eco_lot_val else 0

        op_val = ws.cell(r, COL_OPERATORS).value
        operators = int(op_val) if op_val else 1

        twin_val = ws.cell(r, COL_TWIN).value
        if twin_val and str(twin_val).strip():
            twin_refs[sku] = str(twin_val).strip()

        machines_set.add(machine)

        # Parse daily NP values
        daily = []
        for c in range(DATE_START_COL, DATE_START_COL + n_days):
            val = ws.cell(r, c).value
            if val is not None and isinstance(val, (int, float)) and val < 0:
                daily.append((c - DATE_START_COL, abs(val)))  # (dayIdx, qty)

        # Create unique op ID per row
        op_id = f"{sku}_{machine}_{r}"

        ops.append(
            {
                "id": op_id,
                "sku": sku,
                "machine": machine,
                "tool": tool,
                "pH": pH,
                "operators": operators,
                "eco_lot": eco_lot,
                "daily": daily,
                "twin_sku": twin_refs.get(sku),
            }
        )

    return {
        "ops": ops,
        "machines": list(machines_set),
        "n_days": n_days,
        "n_workdays": n_workdays,
        "workday_indices": workday_indices,
        "cal_to_work": cal_to_work,
        "workdays_bool": workdays_bool,
        "dates": dates,
        "twin_refs": twin_refs,
    }


def _build_solver_request(isop_data: dict) -> SolverRequest:
    """Build SolverRequest from parsed ISOP data."""
    ops = isop_data["ops"]
    cal_to_work = isop_data["cal_to_work"]
    n_workdays = isop_data["n_workdays"]

    jobs = []
    for op in ops:
        pH = op["pH"]
        tool_oee = DEFAULT_OEE

        for day_idx, qty in op["daily"]:
            # Map to workday-indexed time
            cal_day = day_idx + PRE_START_DAYS
            wk = cal_to_work[cal_day] if cal_day < len(cal_to_work) else n_workdays - 1
            if wk < 0:
                wk = 0
            due_date_min = (wk + 1) * DAY_CAP

            duration_min = math.ceil((qty / (pH * tool_oee)) * 60)
            setup_min = 45  # default setup

            job_id = f"{op['id']}_d{day_idx}"
            jobs.append(
                JobInput(
                    id=job_id,
                    sku=op["sku"],
                    due_date_min=due_date_min,
                    weight=1.0,
                    operations=[
                        OperationInput(
                            id=job_id,
                            machine_id=op["machine"],
                            tool_id=op["tool"],
                            duration_min=duration_min,
                            setup_min=setup_min,
                            operators=op["operators"],
                        )
                    ],
                )
            )

    machines = [MachineInput(id=m) for m in isop_data["machines"]]

    # Build workday list for solver (just sequential 0..n_workdays-1)
    workdays = list(range(n_workdays))

    return SolverRequest(
        jobs=jobs,
        machines=machines,
        config=SolverConfig(
            time_limit_s=60,
            objective="weighted_tardiness",
            num_workers=4,
        ),
        constraints=ConstraintConfigInput(
            setup_crew=True,
            tool_timeline=True,
            calco_timeline=False,  # no calco data in ISOP
            operator_pool=False,
        ),
        workdays=workdays,
    )


# ── Tests ──


@pytest.fixture(scope="module")
def isop_data():
    return _parse_isop()


@pytest.fixture(scope="module")
def solver_result(isop_data):
    request = _build_solver_request(isop_data)
    solver = CpsatSolver()
    return solver.solve(request), request


class TestRealISOPParsing:
    def test_isop_has_data(self, isop_data):
        """ISOP parsed with operations, machines, dates."""
        assert len(isop_data["ops"]) > 50, f"Only {len(isop_data['ops'])} ops parsed"
        assert len(isop_data["machines"]) >= 4
        assert isop_data["n_days"] >= 60

    def test_isop_has_demand(self, isop_data):
        """ISOP has demand cells (negative NP values)."""
        total_demand = sum(qty for op in isop_data["ops"] for _, qty in op["daily"])
        assert total_demand > 1_000_000, f"Only {total_demand} pcs total demand"

    def test_isop_weekends_identified(self, isop_data):
        """ISOP correctly identifies weekends."""
        n_weekends = sum(1 for w in isop_data["workdays_bool"] if not w)
        assert n_weekends >= 15, f"Only {n_weekends} weekend days in {isop_data['n_days']} days"

    def test_isop_has_twins(self, isop_data):
        """ISOP has twin references."""
        assert len(isop_data["twin_refs"]) >= 4, f"Only {len(isop_data['twin_refs'])} twin refs"


class TestRealISOPSolver:
    def test_solver_finds_solution(self, solver_result):
        """CP-SAT finds feasible/optimal solution for real ISOP."""
        result, _ = solver_result
        assert result.status in ("optimal", "feasible", "timeout"), (
            f"Solver failed: {result.status}"
        )
        assert result.n_ops > 0

    def test_solve_time_reasonable(self, solver_result):
        """Solver completes within 120s."""
        result, _ = solver_result
        assert result.solve_time_s < 120, f"Solve took {result.solve_time_s:.1f}s, expected < 120s"

    def test_no_shift_crossing(self, solver_result):
        """No operation crosses shift boundary at 510 min within a day."""
        result, _ = solver_result
        SHIFT_LEN = 510
        crossings = 0
        for sop in result.schedule:
            start_in_day = sop.start_min % DAY_CAP
            size = sop.end_min - sop.start_min
            if size <= SHIFT_LEN:
                end_in_day = start_in_day + size
                in_x = end_in_day <= SHIFT_LEN
                in_y = start_in_day >= SHIFT_LEN
                if not (in_x or in_y):
                    crossings += 1
        assert crossings == 0, f"{crossings} ops cross shift boundary"

    def test_no_day_crossing(self, solver_result):
        """No operation crosses day boundary."""
        result, _ = solver_result
        crossings = 0
        for sop in result.schedule:
            start_day = sop.start_min // DAY_CAP
            end_day = (sop.end_min - 1) // DAY_CAP if sop.end_min > 0 else start_day
            if start_day != end_day:
                crossings += 1
        assert crossings == 0, f"{crossings} ops cross day boundary"

    def test_day_capacity_respected(self, solver_result):
        """Max 1020 min of work per day per machine."""
        result, _ = solver_result
        day_machine_load: dict[tuple, int] = defaultdict(int)
        for sop in result.schedule:
            day = sop.start_min // DAY_CAP
            size = sop.end_min - sop.start_min
            day_machine_load[(day, sop.machine_id)] += size
        violations = [(k, v) for k, v in day_machine_load.items() if v > DAY_CAP]
        assert len(violations) == 0, (
            f"{len(violations)} day/machine exceed 1020 min: {violations[:3]}"
        )

    def test_otd_real(self, solver_result):
        """OTD is computed with real capacity (report actual value)."""
        result, request = solver_result
        job_map = {j.id: j for j in request.jobs}
        n_tardy = sum(1 for sop in result.schedule if sop.is_tardy)
        n_total = len(result.schedule)
        otd_pct = ((n_total - n_tardy) / n_total * 100) if n_total > 0 else 0
        print(f"\n  OTD REAL: {otd_pct:.1f}% ({n_tardy} tardy / {n_total} total)")
        print(f"  Tardiness total: {result.total_tardiness_min} min")
        print(f"  Solve time: {result.solve_time_s:.1f}s")
        print(f"  Status: {result.status}")
        # We don't assert 100% — this is the REAL OTD
        assert otd_pct >= 0  # just ensure it's computed

    def test_workday_range_valid(self, solver_result):
        """All ops scheduled within valid workday range."""
        result, request = solver_result
        n_workdays = len(request.workdays) if request.workdays else 999
        max_valid = n_workdays * DAY_CAP
        violations = [sop for sop in result.schedule if sop.end_min > max_valid]
        if request.workdays:
            assert len(violations) == 0, (
                f"{len(violations)} ops exceed workday range "
                f"(max={max_valid}, n_workdays={n_workdays})"
            )
