"""Tests for the Python ISOP parser (F2-02).

Creates minimal XLSX fixtures in-memory to validate:
  - Header detection
  - Column mapping
  - Date parsing
  - NP value parsing + red cell detection
  - NikufraData structure
  - Trust score
  - Error cases
"""

from __future__ import annotations

import io
from datetime import date

import openpyxl
from openpyxl.styles import Font

from src.domain.nikufra.column_mapper import build_column_map as _build_column_map
from src.domain.nikufra.helpers import (
    _is_red_color,
    _normalize_code,
    _parse_date_cell,
    _parse_numeric,
)
from src.domain.nikufra.isop_parser import parse_isop_file

# ── Helpers ──────────────────────────────────────────────────────


def _make_isop_xlsx(
    *,
    header_row: int = 7,
    headers: list[str] | None = None,
    data_rows: list[list] | None = None,
    sheet_name: str = "Planilha1",
    workday_flags_row: int | None = None,
    workday_flags: list[int] | None = None,
    red_cells: list[tuple[int, int]] | None = None,
) -> bytes:
    """Create a minimal ISOP XLSX in memory.

    header_row: 1-based row number for headers.
    headers: Column headers (at least "Referência Artigo" and "Máquina").
    data_rows: Data rows (matching header columns).
    red_cells: List of (row, col) 1-based to highlight red.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if headers is None:
        headers = [
            "Cliente",
            "Nome",
            "Produto Acabado",
            "Referência Artigo",
            "Designação",
            "Lote Económico",
            "Máquina",
            "Máquina alternativa",
            "Ferramenta",
            "Tp.Setup",
            "Peças/H",
            "Nº Pessoas",
            "Atraso",
            "Peça Gémea",
            # Date columns
            date(2026, 3, 2),
            date(2026, 3, 3),
            date(2026, 3, 4),
        ]

    # Write workday flags if provided
    if workday_flags_row and workday_flags:
        for ci, val in enumerate(workday_flags):
            # Offset to match date columns (skip text columns)
            date_start = len([h for h in headers if not isinstance(h, date)])
            ws.cell(row=workday_flags_row, column=date_start + ci + 1, value=val)

    # Write headers
    for ci, h in enumerate(headers):
        cell = ws.cell(row=header_row, column=ci + 1)
        if isinstance(h, date):
            cell.value = h
        else:
            cell.value = h

    # Write data rows
    if data_rows:
        for ri, row in enumerate(data_rows):
            for ci, val in enumerate(row):
                ws.cell(row=header_row + 1 + ri, column=ci + 1, value=val)

    # Apply red highlighting
    if red_cells:
        red_font = Font(color="FFFF0000")
        for r, c in red_cells:
            ws.cell(row=r, column=c).font = red_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Unit tests: helpers ──────────────────────────────────────────


class TestHelpers:
    def test_normalize_code(self):
        assert _normalize_code("  prm019 ") == "PRM019"
        assert _normalize_code(None) == ""
        assert _normalize_code(123) == "123"

    def test_parse_numeric(self):
        assert _parse_numeric(42) == 42.0
        assert _parse_numeric("3,5") == 3.5
        assert _parse_numeric(None, 99) == 99.0
        assert _parse_numeric("abc", 0) == 0.0

    def test_parse_date_cell_datetime(self):
        from datetime import datetime

        d = _parse_date_cell(datetime(2026, 3, 5, 10, 30))
        assert d == date(2026, 3, 5)

    def test_parse_date_cell_iso_string(self):
        assert _parse_date_cell("2026-03-05") == date(2026, 3, 5)

    def test_parse_date_cell_dmy_string(self):
        assert _parse_date_cell("05/03/2026") == date(2026, 3, 5)

    def test_parse_date_cell_dm_string(self):
        d = _parse_date_cell("05/03")
        assert d is not None
        assert d.month == 3
        assert d.day == 5

    def test_parse_date_cell_none(self):
        assert _parse_date_cell(None) is None

    def test_is_red_color(self):
        assert _is_red_color("FF0000") is True
        assert _is_red_color("FFFF0000") is True  # with alpha
        assert _is_red_color("00FF00") is False
        assert _is_red_color("0000FF") is False
        assert _is_red_color("C80000") is True  # dark red
        assert _is_red_color(None) is False


class TestColumnMap:
    def test_basic_mapping(self):
        headers = [
            "Cliente",
            "Nome",
            "Produto Acabado",
            "Referência Artigo",
            "Designação",
            "Lote Económico",
            "Máquina",
            "Máquina alternativa",
            "Ferramenta",
            "Tp.Setup",
            "Peças/H",
            "Nº Pessoas",
            "Atraso",
        ]
        cm = _build_column_map(headers)
        assert cm is not None
        assert cm.ref_artigo == 3
        assert cm.maquina == 6
        assert cm.ferramenta == 8
        assert cm.pecas_h == 10

    def test_missing_required_returns_none(self):
        headers = ["Cliente", "Nome", "Designação"]
        assert _build_column_map(headers) is None

    def test_case_insensitive(self):
        headers = ["REFERÊNCIA ARTIGO", "MÁQUINA", "FERRAMENTA"]
        cm = _build_column_map(headers)
        assert cm is not None
        assert cm.ref_artigo == 0
        assert cm.maquina == 1


# ── Integration tests ────────────────────────────────────────────


class TestParseISOPFile:
    def test_basic_parse(self):
        xlsx = _make_isop_xlsx(
            data_rows=[
                [
                    "CL001",
                    "Cliente A",
                    "PA001",
                    "SKU001",
                    "Peça Teste",
                    1000,
                    "PRM019",
                    "PRM031",
                    "BFP001",
                    1.5,
                    500,
                    2,
                    0,
                    "",
                    -15000,
                    None,
                    -8000,
                ],
            ],
        )
        result = parse_isop_file(io.BytesIO(xlsx))

        assert result.success is True
        assert result.data is not None

        data = result.data
        assert len(data["operations"]) == 1
        assert data["operations"][0]["sku"] == "SKU001"
        assert data["operations"][0]["m"] == "PRM019"
        assert data["operations"][0]["t"] == "BFP001"
        assert data["operations"][0]["pH"] == 500.0
        assert data["operations"][0]["cl"] == "CL001"

        # NP values
        d = data["operations"][0]["d"]
        assert d[0] == -15000
        assert d[1] is None
        assert d[2] == -8000

        # Dates
        assert len(data["dates"]) == 3
        assert data["dates"][0] == "02/03"

        # Machines
        machine_ids = [m["id"] for m in data["machines"]]
        assert "PRM019" in machine_ids
        assert "PRM031" in machine_ids  # alt machine

        # Tools
        assert len(data["tools"]) == 1
        assert data["tools"][0]["id"] == "BFP001"
        assert data["tools"][0]["skus"] == ["SKU001"]

    def test_multiple_rows_same_tool(self):
        """Two SKUs on the same tool → tool.skus has both."""
        xlsx = _make_isop_xlsx(
            data_rows=[
                [
                    "CL001",
                    "A",
                    "PA001",
                    "SKU001",
                    "Peça A",
                    1000,
                    "PRM019",
                    "-",
                    "BFP001",
                    1.0,
                    400,
                    1,
                    0,
                    "SKU002",
                    -5000,
                    None,
                    None,
                ],
                [
                    "CL001",
                    "A",
                    "PA002",
                    "SKU002",
                    "Peça B",
                    1000,
                    "PRM019",
                    "-",
                    "BFP001",
                    1.0,
                    400,
                    1,
                    0,
                    "SKU001",
                    None,
                    -3000,
                    None,
                ],
            ],
        )
        result = parse_isop_file(io.BytesIO(xlsx))
        assert result.success
        data = result.data

        assert len(data["operations"]) == 2
        assert len(data["tools"]) == 1
        assert set(data["tools"][0]["skus"]) == {"SKU001", "SKU002"}

        # Twin fields
        assert data["operations"][0].get("twin") == "SKU002"
        assert data["operations"][1].get("twin") == "SKU001"

    def test_red_cell_np_negation(self):
        """Positive value with red font → negative NP."""
        xlsx = _make_isop_xlsx(
            data_rows=[
                [
                    "CL001",
                    "A",
                    "PA001",
                    "SKU001",
                    "Peça",
                    1000,
                    "PRM019",
                    "-",
                    "BFP001",
                    1.0,
                    400,
                    1,
                    0,
                    "",
                    15000,
                    None,
                    None,
                ],
            ],
            # Red-highlight the first date cell (row 8, col 15)
            red_cells=[(8, 15)],
        )
        result = parse_isop_file(io.BytesIO(xlsx))
        assert result.success
        # Positive red → negated
        assert result.data["operations"][0]["d"][0] == -15000

    def test_missing_sheet(self):
        xlsx = _make_isop_xlsx(sheet_name="WrongSheet")
        result = parse_isop_file(io.BytesIO(xlsx))
        assert not result.success
        assert "Planilha1" in result.errors[0]

    def test_no_header_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planilha1"
        ws.cell(row=1, column=1, value="Random data")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        result = parse_isop_file(buf)
        assert not result.success
        assert "cabeçalho" in result.errors[0].lower()

    def test_no_data_rows(self):
        """Headers present but no data rows → error."""
        xlsx = _make_isop_xlsx(data_rows=[])
        result = parse_isop_file(io.BytesIO(xlsx))
        assert not result.success
        assert "dados" in result.errors[0].lower()

    def test_trust_score(self):
        xlsx = _make_isop_xlsx(
            data_rows=[
                [
                    "CL001",
                    "A",
                    "PA001",
                    "SKU001",
                    "Peça",
                    1000,
                    "PRM019",
                    "-",
                    "BFP001",
                    1.0,
                    400,
                    2,
                    0,
                    "",
                    -5000,
                    None,
                    None,
                ],
            ],
        )
        result = parse_isop_file(io.BytesIO(xlsx))
        assert result.success
        assert result.meta is not None
        assert result.meta["trustScore"] > 0
        assert "completeness" in result.meta["trustDimensions"]

    def test_meta_counts(self):
        xlsx = _make_isop_xlsx(
            data_rows=[
                [
                    "CL001",
                    "A",
                    "PA001",
                    "SKU001",
                    "A",
                    1000,
                    "PRM019",
                    "-",
                    "BFP001",
                    1.0,
                    400,
                    1,
                    0,
                    "",
                    -5000,
                    None,
                    None,
                ],
                [
                    "CL002",
                    "B",
                    "PA002",
                    "SKU002",
                    "B",
                    2000,
                    "PRM031",
                    "-",
                    "BFP002",
                    0.5,
                    300,
                    1,
                    0,
                    "",
                    None,
                    -3000,
                    None,
                ],
            ],
        )
        result = parse_isop_file(io.BytesIO(xlsx))
        assert result.success
        assert result.meta["rows"] == 2
        assert result.meta["skus"] == 2
        assert result.meta["machines"] == 2
        assert result.meta["tools"] == 2
        assert result.meta["dates"] == 3

    def test_workday_flags(self):
        """Workday flags from row above header."""
        xlsx = _make_isop_xlsx(
            workday_flags_row=6,
            workday_flags=[1, 0, 1],
            data_rows=[
                [
                    "CL001",
                    "A",
                    "PA001",
                    "SKU001",
                    "A",
                    1000,
                    "PRM019",
                    "-",
                    "BFP001",
                    1.0,
                    400,
                    1,
                    0,
                    "",
                    -5000,
                    None,
                    None,
                ],
            ],
        )
        result = parse_isop_file(io.BytesIO(xlsx))
        assert result.success
        # Should detect workday flags
        assert result.data is not None
        assert len(result.data["workday_flags"]) == 3

    def test_machine_area_assignment(self):
        xlsx = _make_isop_xlsx(
            data_rows=[
                [
                    "CL001",
                    "A",
                    "PA001",
                    "SKU001",
                    "A",
                    1000,
                    "PRM042",
                    "-",
                    "BFP001",
                    1.0,
                    400,
                    1,
                    0,
                    "",
                    -5000,
                    None,
                    None,
                ],
            ],
        )
        result = parse_isop_file(io.BytesIO(xlsx))
        assert result.success
        m = [m for m in result.data["machines"] if m["id"] == "PRM042"][0]
        assert m["area"] == "PG2"

    def test_customers_deduplicated(self):
        xlsx = _make_isop_xlsx(
            data_rows=[
                [
                    "CL001",
                    "Client A",
                    "PA001",
                    "SKU001",
                    "A",
                    1000,
                    "PRM019",
                    "-",
                    "BFP001",
                    1.0,
                    400,
                    1,
                    0,
                    "",
                    -5000,
                    None,
                    None,
                ],
                [
                    "CL001",
                    "Client A",
                    "PA002",
                    "SKU002",
                    "B",
                    2000,
                    "PRM019",
                    "-",
                    "BFP002",
                    0.5,
                    300,
                    1,
                    0,
                    "",
                    None,
                    -3000,
                    None,
                ],
            ],
        )
        result = parse_isop_file(io.BytesIO(xlsx))
        assert result.success
        assert len(result.data["customers"]) == 1
        assert result.data["customers"][0]["code"] == "CL001"
